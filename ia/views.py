# ia/views.py

import json
import types
import pandas as pd 
import logging
import traceback
import tempfile
import re # Importar o módulo de regex
import base64
import mimetypes
import io # Para trabalhar com streams de bytes na extração
from pathlib import Path
from django.utils import timezone
import google.generativeai as genai
import typing
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from django import forms
from .models import Chat, ChatMessage, ChatAttachment, ApiUsageLog
from ia.retrieval import find_relevant_document_chunks
from ia.old_spreadsheet_utils import (
    convert_to_csv,
    load_dataframe,
    normalize_columns,
    parse_filters_from_prompt,
    apply_filters,
    df_to_prompt_block,
    SpreadsheetExtractionError,
)

from ia.openai_helper import get_or_create_thread, ensure_oa_file, oa_client
# --- Imports das bibliotecas de extração ---
# Guardados por try-except para não quebrar se não instalados
try:
    import docx
except ImportError:
    docx = None
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import fitz  # PyMuPDF para extração de PDF
    pymupdf_available = True
except ImportError:
    fitz = None
    pymupdf_available = False
try:
    # Importa as partes necessárias do odfpy
    import odf.opendocument
    import odf.text
    import odf.teletype # Necessário para extractText
    odfpy_available = True
except ImportError:
    odfpy_available = False # Define como False se não estiver instalado

logger = logging.getLogger(__name__)

# --- Configurações e Constantes ---
MAX_UPLOAD_SIZE_MB = getattr(settings, 'MAX_CHAT_UPLOAD_MB', 15)
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024
# Limite de caracteres para extração de texto de arquivos
MAX_EXTRACTION_CHARS = getattr(settings, 'IA_MAX_FILE_EXTRACTION_CHARS', 1000000)
EMOJI_PATTERN = r"[\U0001F300-\U0001FAFF\U00002700-\U000027B0]"

# Modelos Gemini Mapeados
MODEL_MAP = {
    # 1.5
    "gemini-1.5-pro":   "models/gemini-1.5-pro-latest",
    "gemini-1.5-flash": "models/gemini-1.5-flash-latest",

    # 2.0 (flash & flash-lite) 
    "gemini-2.0-flash":      "models/gemini-2.0-flash",
    "gemini-2.0-flash-lite": "models/gemini-2.0-flash-lite",

    # 2.5 preview
    "gemini-2.5-pro":   "models/gemini-2.5-pro-preview-03-25",
    "gemini-2.5-flash": "models/gemini-2.5-flash-preview-04-17",

    #GPT
    "gpt-4o-mini": "gpt-4o-mini",
    "gpt-4o": "gpt-4o", 
}

MODEL_MAP.update({
    # ‘roto-ia’ usa internamente o mesmo Gemini que já é rápido/económico.
    # Se quiser pode pôr 1.5-flash, 2.0-flash, 2.5-flash… aqui.
    "roto-ia": MODEL_MAP["gemini-2.5-flash"],
})

DEFAULT_GEMINI_MODEL_KEY = "gemini-1.5-flash"
DEFAULT_GEMINI_MODEL_NAME = MODEL_MAP.get(DEFAULT_GEMINI_MODEL_KEY, "models/gemini-1.5-flash-latest")

# Extensões suportadas para extração de texto
EXTRACTABLE_EXTENSIONS = {
    '.pdf', '.docx', '.xlsx', '.odt', '.txt', '.csv', '.md', '.rtf',
    '.py', '.js', '.html', '.css'
}
# Mimetypes de imagem suportados pela API Gemini
SUPPORTED_IMAGE_MIMETYPES_FOR_API = [
    'image/jpeg', 'image/png', 'image/webp', 'image/gif', 'image/heic', 'image/heif'
]

# Importa a função de cálculo de custo e formulários
try:
    from .cost_calculator import calculate_gemini_cost
except ImportError:
    logger.error("Arquivo ia/cost_calculator.py não encontrado. Cálculo de custo desativado.")
    def calculate_gemini_cost(*args, **kwargs): return Decimal(0)
try:
    from .forms import ApiUsageFilterForm
except ImportError:
    logger.warning("Arquivo ia/forms.py ou ApiUsageFilterForm não encontrado.")
    ApiUsageFilterForm = None

from django.contrib.auth import get_user_model
User = get_user_model()

# --- Configuração da API Gemini ---
try:
    if hasattr(settings, 'GEMINI_API_KEY') and settings.GEMINI_API_KEY:
        genai.configure(api_key=settings.GEMINI_API_KEY)
        logger.info("API Key do Gemini configurada com sucesso.")
    else:
        logger.warning("GEMINI_API_KEY não encontrada ou vazia nas configurações.")
except Exception as e:
    logger.error(f"Erro ao configurar a API Key do Gemini: {e}")


# --- Funções Auxiliares ---

def remove_emojis(text):
    """Remove emojis de uma string."""
    if not text or not isinstance(text, str):
        return text
    return re.sub(EMOJI_PATTERN, '', text).strip()

def extract_text_from_file(filename, file_content_bytes):
    """
    Tenta extrair texto de diferentes tipos de arquivo usando bibliotecas apropriadas.
    Retorna o texto extraído (limitado em tamanho) ou uma string de erro/aviso.
    """
    logger.debug(f"FUNÇÃO extract_text_from_file INICIADA para: '{filename}'")
    logger.debug(f"Tentando extrair texto de '{filename}' (Tamanho: {len(file_content_bytes)} bytes)")
    ext = '.' + filename.split('.')[-1].lower() if '.' in filename else ''
    text = None
    library_missing = False
    error_message = None

    try:
        # --- DOCX ---
        if ext == '.docx':
            if not docx:
                library_missing = True
                raise ImportError("python-docx não está instalado.")
            document = docx.Document(io.BytesIO(file_content_bytes))
            text = "\n".join([para.text for para in document.paragraphs if para.text])

        # --- XLSX (sem limite de linhas/colunas) ---
        elif ext == '.xlsx':
            if not openpyxl:
                library_missing = True
                raise ImportError("openpyxl não está instalado.")
            workbook = openpyxl.load_workbook(
                filename=io.BytesIO(file_content_bytes),
                data_only=True
            )
            all_texts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                n_rows = sheet.max_row
                n_cols = sheet.max_column
                all_texts.append(f"--- Planilha: {sheet_name} ({n_rows} linhas x {n_cols} colunas) ---")
                for row in sheet.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols):
                    row_vals = [
                        str(cell.value)
                        for cell in row
                        if cell.value is not None and str(cell.value).strip()
                    ]
                    if row_vals:
                        all_texts.append(" | ".join(row_vals))
            text = "\n".join(all_texts)

        # --- PDF (PyMuPDF) ---
        elif ext == '.pdf':
            logger.debug(f"Processando PDF com PyMuPDF. Biblioteca disponível: {pymupdf_available}")
            if not pymupdf_available:
                library_missing = True
                raise ImportError("PyMuPDF (fitz) não está instalado.")
            doc = fitz.open(stream=file_content_bytes, filetype="pdf")
            num_pages = len(doc)
            all_texts = []
            max_pages = 200
            for i, page in enumerate(doc.pages()):
                if i >= max_pages:
                    all_texts.append(f"\n(...mais {num_pages - max_pages} páginas não incluídas...)")
                    break
                page_text = page.get_text("text", sort=True)
                if page_text:
                    all_texts.append(page_text)
            doc.close()
            text = "\n".join(all_texts)
            if not text:
                error_message = f"(Não foi possível extrair conteúdo textual de '{filename}')"

        # --- XLS (xlrd stub) ---
        elif ext == ".xls":
            import xlrd
            workbook = xlrd.open_workbook(file_contents=file_content_bytes)
            # implementar extração conforme necessário
            text = "(Extração de .xls não implementada)"

        # --- ODS (pyexcel_ods stub) ---
        elif ext == ".ods":
            import pyexcel_ods
            data = pyexcel_ods.get_data(io.BytesIO(file_content_bytes))
            # você pode formatar 'data' conforme desejar
            text = str(data)

        # --- ODT ---
        elif ext == '.odt':
            if not odfpy_available:
                library_missing = True
                raise ImportError("odfpy não está instalado.")
            doc = odf.opendocument.load(io.BytesIO(file_content_bytes))
            all_texts = []
            text_elements = doc.getElementsByType(odf.text.P)
            for elem in text_elements:
                para_text = odf.teletype.extractText(elem)
                if para_text:
                    all_texts.append(para_text.strip())
            text = "\n".join(all_texts)
            if not text:
                error_message = f"(Não foi possível extrair conteúdo textual de '{filename}')"

        # --- Texto simples (txt, csv, md, etc.) ---
        elif ext in EXTRACTABLE_EXTENSIONS:
            try:
                text = file_content_bytes.decode('utf-8')
            except UnicodeDecodeError:
                logger.warning(f"Falha ao decodificar '{filename}' como UTF-8, tentando latin-1.")
                text = file_content_bytes.decode('latin-1', errors='replace')

        # --- Extensão não suportada ---
        else:
            logger.warning(f"Extração de texto não suportada para extensão '{ext}' do arquivo '{filename}'.")
            return None

        # --- Processamento final do texto extraído ---
        if text:
            original_len = len(text)
            if original_len > MAX_EXTRACTION_CHARS:
                text = text[:MAX_EXTRACTION_CHARS] + f"\n\n(... [CONTEÚDO TRUNCADO em {MAX_EXTRACTION_CHARS} caracteres] ...)"
                logger.info(
                    f"Texto extraído de '{filename}' (ext: {ext}) TRUNCADO para "
                    f"{MAX_EXTRACTION_CHARS} caracteres (original: {original_len})."
                )
            else:
                logger.info(
                    f"Texto extraído com sucesso de '{filename}' (ext: {ext}, {original_len} caracteres)."
                )
            # Remove linhas vazias e espaços extras
            text = "\n".join(line.strip() for line in text.splitlines() if line.strip())
            return text

        elif error_message:
            return error_message

        elif not library_missing and ext in ['.docx', '.xlsx', '.pdf', '.odt']:
            logger.warning(
                f"Biblioteca para '{ext}' foi encontrada, mas não retornou texto de '{filename}'."
            )
            return f"(Não foi possível extrair conteúdo textual de '{filename}')"

        else:
            return None

    except ImportError as ie:
        logger.error(
            f"Erro de importação ao tentar processar '{filename}' (ext: {ext}): {ie}. "
            "Verifique se a biblioteca está instalada."
        )
        return f"(Erro: Biblioteca para processar arquivos '{ext}' não encontrada no servidor.)"

    except Exception as e:
        logger.error(
            f"Erro inesperado ao extrair texto de '{filename}' (ext: {ext}): {e}",
            exc_info=True
        )
        return f"(Erro ao tentar ler o conteúdo de '{filename}')"

def log_api_usage(user, model_name, ai_message, usage_metadata, image_count):
    """Tenta registrar o uso da API e calcular o custo em USD e BRL."""
    if not usage_metadata:
        logger.warning(f"Metadados de uso não encontrados para msg {ai_message.id if ai_message else '?'}. Custo não registrado.")
        return

    try:
        input_tokens = getattr(usage_metadata, 'prompt_token_count', 0)
        output_tokens = getattr(usage_metadata, 'candidates_token_count', 0)

        # Calcula o custo estimado em USD
        estimated_cost_usd = calculate_gemini_cost(
            model_name=model_name,
            input_tokens=input_tokens,
            output_tokens=output_tokens,
            image_count=image_count
        )

        # Converte para BRL
        usd_to_brl_rate = getattr(settings, 'USD_TO_BRL_RATE', None)
        estimated_cost_brl = Decimal('0.0')

        if usd_to_brl_rate:
             try:
                 rate = Decimal(str(usd_to_brl_rate))
                 if rate > 0:
                     estimated_cost_brl = estimated_cost_usd * rate
                 else:
                      logger.error(f"USD_TO_BRL_RATE ('{usd_to_brl_rate}') inválida (não positiva). Custo BRL não calculado.")
             except (ValueError, TypeError):
                  logger.error(f"USD_TO_BRL_RATE ('{usd_to_brl_rate}') inválida (não é número). Custo BRL não calculado.")
        else:
             logger.warning("USD_TO_BRL_RATE não definida nas configurações. Custo BRL não calculado.")


        # Cria e salva o registro de log
        log_entry = ApiUsageLog.objects.create(
            user=user,
            model_name=model_name,
            ai_message=ai_message,
            input_tokens=input_tokens,
            output_tokens=output_tokens,
            image_count=image_count,
            estimated_cost=estimated_cost_usd,
            estimated_cost_brl=estimated_cost_brl,
            timestamp=ai_message.created_at if ai_message else timezone.now()
        )
        logger.info(f"Uso API registrado: Log {log_entry.id}, User: {user.username}, Model: {model_name}, "
                    f"Tokens In: {input_tokens}, Out: {output_tokens}, Imgs: {image_count}, "
                    f"Cost: ${estimated_cost_usd:.6f} / R${estimated_cost_brl:.6f}")

    except AttributeError as ae:
        logger.error(f"Erro ao acessar atributos usage_metadata para msg {ai_message.id if ai_message else '?'}: {ae}. Metadados: {usage_metadata}", exc_info=True)
    except Exception as e:
        logger.error(f"Erro ao registrar log de uso API para msg {ai_message.id if ai_message else '?'}: {e}", exc_info=True)


# --- Views ---

@login_required
def chat_page_view(request):
    """Renderiza a página do chat (HTML)."""
    api_key_configured = bool(hasattr(settings, 'GEMINI_API_KEY') and settings.GEMINI_API_KEY)
    context = {
        'api_key_configured': api_key_configured
    }
    return render(request, 'ia/chat_page.html', context)

@login_required
@csrf_exempt
@require_http_methods(["GET"])
def list_chats_view(request):
    """Lista todas as conversas do usuário logado, garantindo retorno JSON."""
    user = request.user
    try:
        chats = Chat.objects.filter(user=user).order_by('-updated_at')
        data = [{
                 'id': chat.id,
                 'title': str(chat.title) if chat.title else f"Chat {chat.id}",
                 'created_at': chat.created_at.isoformat() if chat.created_at else None,
                 'updated_at': chat.updated_at.isoformat() if chat.updated_at else None,
             } for chat in chats]
        return JsonResponse(data, safe=False)
    except Exception as e:
        logger.error(f"Erro ao listar chats para {user.username}: {e}", exc_info=True)
        return JsonResponse({'error': 'Erro interno ao buscar lista de chats.'}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_chat_view(request):
    """Cria uma nova conversa para o usuário logado."""
    user = request.user
    try:
        title = 'Nova Conversa'
        new_chat = Chat.objects.create(user=user, title=title)
        logger.info(f"Chat {new_chat.id} ('{new_chat.title}') criado para usuário {user.username}.")
        return JsonResponse({
            'id': new_chat.id,
            'title': new_chat.title,
            'created_at': new_chat.created_at.isoformat(),
            'updated_at': new_chat.updated_at.isoformat(),
        }, status=201)
    except Exception as e:
        logger.exception(f"Erro ao criar nova conversa para {user.username}: {e}")
        return JsonResponse({'error': 'Erro interno ao criar nova conversa.'}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["GET"])
def get_chat_view(request, chat_id):
    """Retorna os detalhes de um chat específico do usuário, incluindo mensagens e anexos."""
    user = request.user
    try:
        chat = get_object_or_404(Chat, id=chat_id, user=user)
        messages_qs = chat.messages.order_by('created_at').prefetch_related('attachments')

        messages_data = []
        for msg in messages_qs:
            msg_attachments_data = []
            for att in msg.attachments.all():
                try:
                    msg_attachments_data.append({
                        'id': att.id,
                        'url': att.file.url,
                        'filename': att.get_filename()
                    })
                except Exception as att_err:
                    logger.error(f"Erro ao processar anexo {att.id} para msg {msg.id} no chat {chat_id}: {att_err}")

            messages_data.append({
                'id': msg.id,
                'sender': msg.sender,
                'text': msg.text,
                'created_at': msg.created_at.isoformat(),
                'attachments': msg_attachments_data
            })

        return JsonResponse({
            'id': chat.id,
            'title': chat.title,
            'messages': messages_data,
            'created_at': chat.created_at.isoformat(),
            'updated_at': chat.updated_at.isoformat(),
        })
    except Chat.DoesNotExist:
        return JsonResponse({'error': 'Chat não encontrado ou não pertence a este usuário.'}, status=404)
    except Exception as e:
        logger.exception(f"Erro ao buscar chat {chat_id} para user {user.username}: {e}")
        return JsonResponse({'error': 'Erro interno ao buscar dados do chat.'}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["DELETE"])
def delete_chat_view(request, chat_id):
    """Exclui uma conversa inteira do usuário logado."""
    user = request.user
    try:
        chat = get_object_or_404(Chat, id=chat_id, user=user)
        chat_title = chat.title
        chat_id_log = chat.id
        chat.delete()
        logger.info(f"Chat {chat_id_log} ('{chat_title}') excluído pelo usuário {user.username}.")
        return JsonResponse({'status': 'deleted', 'message': f'Chat "{chat_title}" excluído com sucesso.'})
    except Chat.DoesNotExist:
        return JsonResponse({'error': 'Chat não encontrado ou não pertence a este usuário.'}, status=404)
    except Exception as e:
        logger.exception(f"Erro ao excluir chat {chat_id} pelo usuário {user.username}: {e}")
        return JsonResponse({'error': 'Erro interno ao excluir o chat.'}, status=500)


# ==============================================================================
# == VIEW COMBINADA: Enviar Mensagem (Texto e/ou Arquivos) para a IA ==
# ==============================================================================
def _merge_plans(last_plan: dict | None, new_plan: dict) -> dict:
    """
    Função helper para fundir o plano da pergunta anterior com o da nova.
    A regra principal é herdar os filtros se a nova pergunta não especificar nenhum.
    """
    if not last_plan or not isinstance(last_plan, dict):
        return new_plan # Se não há plano anterior, o novo plano é o plano final.

    final_plan = new_plan.copy()
    
    # Se o novo plano não tem filtros, mas o plano anterior tinha, herda os filtros.
    if not final_plan.get("filter") and last_plan.get("filter"):
        logger.info(f"[PLAN MERGE] Herdou filtros do plano anterior: {last_plan.get('filter')}")
        final_plan["filter"] = last_plan["filter"]
        # Se herdou filtros, faz sentido herdar a ordenação também,
        # a menos que o novo plano explicitamente a remova (o que não é um caso comum).
        if not final_plan.get("sort") and last_plan.get("sort"):
             logger.info(f"[PLAN MERGE] Herdou ordenação do plano anterior: {last_plan.get('sort')}")
             final_plan["sort"] = last_plan["sort"]
             
    return final_plan


# --- Sua view send_message_view ---

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def send_message_view(request, chat_id):
    """Envia prompt + anexos (opcional) ao Assistant, aguarda execução e armazena a resposta."""

    user = request.user
    chat = get_object_or_404(Chat, id=chat_id, user=user)

    prompt: str = (request.POST.get("prompt") or "").strip()
    files = request.FILES.getlist("arquivo")

    if not prompt and not files:
        return JsonResponse({"error": "Envie um prompt ou um arquivo."}, status=400)

    # ------------------------------------------------------------------
    # 1) Registra mensagem do usuário
    # ------------------------------------------------------------------
    user_msg = ChatMessage.objects.create(
        chat=chat,
        sender="user",
        text=prompt or "(envio de arquivo)"
    )

    # ------------------------------------------------------------------
    # 2) Salva anexos – se houver – e prepara lista para OpenAI
    # ------------------------------------------------------------------
    dj_attachments: list[ChatAttachment] = []
    for f in files:
        dj_attachments.append(
            ChatAttachment.objects.create(chat=chat, message=user_msg, file=f)
        )

    try:
        # ------------------------------------------------------------------
        # 3) Garante thread e monta mensagem
        # ------------------------------------------------------------------
        thread_id = get_or_create_thread(chat)

        msg_kwargs = {
            "role": "user",
            "content": prompt or "Descreva os dados enviados e gere análises.",
        }

        if dj_attachments:
            msg_kwargs["attachments"] = [
                {
                    "file_id": ensure_oa_file(att),
                    "tools": [{"type": "code_interpreter"}],
                }
                for att in dj_attachments
            ]

        oa_client.beta.threads.messages.create(thread_id, **msg_kwargs)

        # ------------------------------------------------------------------
        # 4) Executa o Assistant (blocking)
        # ------------------------------------------------------------------
        run = oa_client.beta.threads.runs.create_and_poll(
            thread_id=thread_id,
            assistant_id=settings.ASSISTANT_ID,
        )

        # ------------------------------------------------------------------
        # 5) Recupera última resposta
        # ------------------------------------------------------------------
        last_msg = oa_client.beta.threads.messages.list(
            thread_id, order="desc", limit=1
        ).data[0]

        def _to_str(part) -> str:
            """Converte cada Fragmento retornado pela API em texto limpo."""
            if isinstance(part, str):
                return part
            if hasattr(part, "text"):
                txt_obj = part.text
                return getattr(txt_obj, "value", str(txt_obj))
            if hasattr(part, "value"):
                return str(part.value)
            return str(part)

        ai_text: str = "".join(_to_str(p) for p in last_msg.content).strip() or (
            "(A IA não retornou texto.)"
        )
        ai_text = remove_emojis(ai_text)

        # ------------------------------------------------------------------
        # 6) Grava mensagem da IA
        # ------------------------------------------------------------------
        ai_msg = ChatMessage.objects.create(chat=chat, sender="ai", text=ai_text)

        # ------------------------------------------------------------------
        # 7) Log de uso (tokens / imagens)
        # ------------------------------------------------------------------
        if getattr(run, "usage", None):
            usage_obj = run.usage
            # A API usa input_tokens / output_tokens nos objetos Run
            in_tok = getattr(usage_obj, "input_tokens", 0)
            out_tok = getattr(usage_obj, "output_tokens", 0)

            ApiUsageLog.objects.create(
                user=user,
                model_name=run.model,
                ai_message=ai_msg,
                input_tokens=in_tok,
                output_tokens=out_tok,
                image_count=len(dj_attachments),
                estimated_cost=Decimal(0),  # calcule se desejar
                estimated_cost_brl=Decimal(0),
                timestamp=ai_msg.created_at,
            )

        # ------------------------------------------------------------------
        # 8) Atualiza conversa e devolve resposta
        # ------------------------------------------------------------------
        chat.updated_at = ai_msg.created_at
        chat.save(update_fields=["updated_at"])

        return JsonResponse(
            {
                "response": ai_text,
                "ai_message_id": ai_msg.id,
                "user_message_id": user_msg.id,
                "uploaded_files": [
                    {
                        "id": att.id,
                        "filename": att.get_filename(),
                        "url": att.file.url,
                    }
                    for att in dj_attachments
                ],
            }
        )

    except Exception as exc:  # noqa: BLE001
        logger.error("Erro no Assistant: %s", exc, exc_info=True)
        chat.updated_at = timezone.now()
        chat.save(update_fields=["updated_at"])
        return JsonResponse({"error": "Falha ao processar sua solicitação."}, status=500)



@login_required
@csrf_exempt
@require_http_methods(["POST"])
def edit_user_message_view(request, chat_id, message_id):
    """
    Edita uma mensagem DO USUÁRIO (apenas texto), remove mensagens posteriores,
    reprocessa a IA (SEM re-extrair anexos), registra o uso e retorna status.
    NOTA: A edição de mensagens que *originalmente* tinham anexos NÃO é suportada.
    """
    user = request.user
    try:
        chat = get_object_or_404(Chat, id=chat_id, user=user)
    except Chat.DoesNotExist:
        return JsonResponse({'error': 'Chat não encontrado ou acesso não permitido.'}, status=404)

    try:
        body = json.loads(request.body)
        new_text_raw = body.get('new_text')
        if not new_text_raw or not isinstance(new_text_raw, str) or not new_text_raw.strip():
            return JsonResponse({'error': 'O novo texto ("new_text") é obrigatório.'}, status=400)

        # Limpa texto e valida
        new_text = remove_emojis(new_text_raw.strip())
        if not new_text:
            return JsonResponse({'error': 'O texto editado não pode ficar vazio ou conter apenas emojis.'}, status=400)

        # Encontra a mensagem e verifica se tem anexos
        msg_to_edit = get_object_or_404(ChatMessage, id=message_id, chat=chat, sender='user')

        # RESTRIÇÃO: Não permite editar mensagens com anexos
        if msg_to_edit.attachments.exists():
             logger.warning(f"Tentativa de editar msg {message_id} com anexos (Chat {chat.id}). Não suportado.")
             return JsonResponse({'error': 'Não é possível editar mensagens que contenham anexos.'}, status=400)

        # Verifica se houve mudança
        if msg_to_edit.text == new_text:
            logger.info(f"Edição da msg {message_id} (Chat {chat.id}) cancelada: texto idêntico.")
            return JsonResponse({'status': 'no_change', 'message': 'Nenhuma alteração necessária.'})

        original_created_at = msg_to_edit.created_at

        # Atualiza o texto da mensagem do usuário
        msg_to_edit.text = new_text
        msg_to_edit.save(update_fields=['text'])
        logger.info(f"Usuário {user.username} editou msg {message_id} (novo texto: '{new_text[:50]}...') no chat {chat_id}.")

        # Remove mensagens posteriores
        subsequent_msgs = chat.messages.filter(created_at__gt=original_created_at)
        deleted_count, deleted_details = subsequent_msgs.delete()
        if deleted_count > 0:
            logger.info(f"{deleted_count} mensagens posteriores ({deleted_details}) deletadas após edição da msg {message_id} no chat {chat_id}.")

        # Monta histórico para reprocessamento (até a msg editada, SEM anexos)
        current_history = []
        for msg in chat.messages.filter(created_at__lte=original_created_at).order_by('created_at'):
            current_history.append({
                'role': 'user' if msg.sender == 'user' else 'model',
                'parts': [{'text': msg.text or ""}] # Apenas texto
            })

        # Adiciona instrução de idioma e anti-emoji
        if current_history and current_history[-1]["role"] == "user":
            instruction = "\n\n(Instruções Importantes: Responda sempre em português brasileiro. NÃO use emojis.)"
            try:
                if current_history[-1].get("parts") and isinstance(current_history[-1]["parts"], list) and len(current_history[-1]["parts"]) > 0:
                    last_part_index = -1
                    for i in range(len(current_history[-1]["parts"]) -1, -1, -1):
                         if "text" in current_history[-1]["parts"][i]: last_part_index = i; break
                    if last_part_index != -1:
                        current_history[-1]["parts"][last_part_index]["text"] = (current_history[-1]["parts"][last_part_index].get("text", "") or "") + instruction
                    else: current_history[-1]["parts"].append({"text": instruction.strip()})
                else: current_history[-1]["parts"] = [{"text": instruction.strip()}]
                logger.debug(f"Instruções adicionadas ao histórico pós-edição (Chat {chat.id}).")
            except Exception as instr_err:
                logger.error(f"Erro ao adicionar instruções pós-edição (Chat {chat.id}): {instr_err}", exc_info=True)

        # Seleciona modelo e Reprocessa com a IA
        selected_model_key = body.get('model', DEFAULT_GEMINI_MODEL_KEY)
        model_name = MODEL_MAP.get(selected_model_key, DEFAULT_GEMINI_MODEL_NAME)
        logger.debug(f"Reprocessando com modelo {model_name} para chat {chat_id} após edição da msg {message_id}.")

        new_ai_response_text = "(Erro ao reprocessar IA após edição)"
        new_ai_message = None
        gemini_resp_edit = None

        try:
            model = genai.GenerativeModel(model_name)
            generation_config = genai.types.GenerationConfig(temperature=0.7)
            safety_settings = [
                 {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
                 {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
                 {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
                 {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
             ]
            # <<< CHAMA A API (Apenas com histórico de texto) >>>
            gemini_resp_edit = model.generate_content(
                current_history,
                generation_config=generation_config,
                safety_settings=safety_settings
            )

            # Processa resposta da IA (similar a send_message_view)
            new_ai_response_text_raw = ""
            try:
                if hasattr(gemini_resp_edit, 'text') and gemini_resp_edit.text:
                    new_ai_response_text_raw = gemini_resp_edit.text.strip()
                elif hasattr(gemini_resp_edit, 'candidates') and gemini_resp_edit.candidates and hasattr(gemini_resp_edit.candidates[0], 'content') and gemini_resp_edit.candidates[0].content.parts:
                     new_ai_response_text_raw = "".join(part.text for part in gemini_resp_edit.candidates[0].content.parts if hasattr(part, 'text')).strip()
                     if not new_ai_response_text_raw:
                         if gemini_resp_edit.prompt_feedback and gemini_resp_edit.prompt_feedback.block_reason:
                              block_reason = gemini_resp_edit.prompt_feedback.block_reason.name
                              new_ai_response_text_raw = f"(A nova resposta foi bloqueada: {block_reason})"
                         else: new_ai_response_text_raw = "(A IA retornou uma resposta vazia após edição.)"
                elif gemini_resp_edit.prompt_feedback and gemini_resp_edit.prompt_feedback.block_reason:
                    block_reason = gemini_resp_edit.prompt_feedback.block_reason.name
                    new_ai_response_text_raw = f"(A nova resposta foi bloqueada: {block_reason})"
                else:
                    new_ai_response_text_raw = "(A IA retornou uma resposta vazia após edição.)"

                new_ai_response_text = remove_emojis(new_ai_response_text_raw)
                if not new_ai_response_text and new_ai_response_text_raw: new_ai_response_text = "(Resposta da IA continha apenas emojis.)"
                elif not new_ai_response_text: new_ai_response_text = "(A IA retornou uma resposta vazia após edição.)"

            except ValueError as ve: # Bloqueio
                new_ai_response_text_raw = f"(A nova resposta foi bloqueada: {ve})"
                new_ai_response_text = new_ai_response_text_raw
            except Exception as resp_err:
                 logger.error(f"Erro ao processar resposta da IA pós-edição (Chat {chat.id}): {resp_err}", exc_info=True)
                 new_ai_response_text_raw = "(Erro interno ao obter nova resposta da IA)"
                 new_ai_response_text = new_ai_response_text_raw


            # Salva a nova resposta da IA
            new_ai_message = ChatMessage.objects.create(chat=chat, sender='ai', text=new_ai_response_text)
            logger.info(f"Nova resposta da IA (msg {new_ai_message.id}) gerada após edição no chat {chat.id}.")

            # LOGAR USO DA API (EDIÇÃO)
            if gemini_resp_edit and hasattr(gemini_resp_edit, 'usage_metadata'):
                log_api_usage(
                    user=user, model_name=model_name, ai_message=new_ai_message,
                    usage_metadata=gemini_resp_edit.usage_metadata, image_count=0 # Edição não envia imagens
                )
            elif gemini_resp_edit:
                 logger.warning(f"Resposta Gemini pós-edição (msg {new_ai_message.id}) não continha 'usage_metadata'. Custo não registrado.")

            # Atualiza o 'updated_at' do chat
            chat.updated_at = new_ai_message.created_at
            chat.save(update_fields=['updated_at'])

            # Retorna sucesso
            return JsonResponse({
                'status': 'edited_and_reprocessed',
                'message': 'Mensagem editada e IA reprocessada com sucesso.',
            })

        except Exception as ia_ex:
            # Erro na chamada da API durante reprocessamento
            logger.error(f"Falha ao reprocessar IA após edição da msg {message_id} no chat {chat_id}: {ia_ex}", exc_info=True)
            # A edição do texto do usuário FOI salva. Atualiza updated_at.
            try:
                chat.updated_at = timezone.now()
                chat.save(update_fields=['updated_at'])
            except Exception as save_err:
                 logger.error(f"Erro adicional ao salvar updated_at após falha no reprocessamento: {save_err}")

            # Retorna status 207 (Multi-Status): edição OK, reprocessamento falhou
            return JsonResponse({
                   'status': 'edited_only',
                   'warning': f'Mensagem editada com sucesso, mas falha ao gerar nova resposta da IA: {type(ia_ex).__name__}'
            }, status=207)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido no corpo da requisição.'}, status=400)
    except ChatMessage.DoesNotExist:
        return JsonResponse({'error': 'Mensagem a editar não encontrada ou inválida.'}, status=404)
    except Exception as e:
        logger.exception(f"Erro inesperado ao editar mensagem {message_id} no chat {chat_id}: {e}")
        return JsonResponse({'error': 'Erro interno ao processar a edição da mensagem.'}, status=500)


# ==============================================================================
# == VIEW DE UPLOAD SEPARADA (LEGADO/REDUNDANTE) ==
# ==============================================================================
@login_required
@csrf_exempt
@require_http_methods(["POST"])
def upload_file_to_chat_view(request, chat_id):
    """(LEGADO?) Recebe arquivos e salva como anexo no chat, SEM interação com IA."""
    user = request.user
    logger.warning(f"Chamada à view LEGADA 'upload_file_to_chat_view' para chat {chat_id} por '{user.username}'.")

    try:
        chat = get_object_or_404(Chat, id=chat_id, user=user)
    except Chat.DoesNotExist:
        return JsonResponse({'error': 'Chat não encontrado ou acesso não permitido.'}, status=404)
    except Exception as e:
        logger.exception(f"Erro ao buscar chat {chat_id} para upload LEGADO: {e}")
        return JsonResponse({'error': 'Erro interno ao buscar dados do chat.'}, status=500)

    if not request.FILES:
        return JsonResponse({'error': 'Nenhum arquivo enviado.'}, status=400)

    files = request.FILES.getlist('arquivo')
    if not files:
        return JsonResponse({'error': 'Nenhum arquivo encontrado no campo "arquivo".'}, status=400)

    logger.info(f"Recebidos {len(files)} arquivo(s) para upload LEGADO no chat {chat.id}.")
    saved_attachments_info = []
    errors = []
    last_upload_time = None

    for f in files:
        original_filename = f.name
        attachment = None
        try:
            logger.debug(f"Processando arquivo '{original_filename}' para upload LEGADO (Chat {chat.id})")
            # Validação de Tamanho
            if f.size > MAX_UPLOAD_SIZE_BYTES:
                raise ValueError(f"Arquivo excede o limite de {MAX_UPLOAD_SIZE_MB}MB.")

            # Salva o anexo SEM mensagem associada
            attachment = ChatAttachment(chat=chat, file=f, message=None)
            attachment.save()
            last_upload_time = attachment.uploaded_at

            saved_filename = attachment.file.name
            display_filename = attachment.get_filename()
            logger.info(f"Anexo '{original_filename}' salvo como '{saved_filename}' (ID: {attachment.id}) no upload LEGADO (Chat {chat.id}).")
            saved_attachments_info.append({
                'id': attachment.id, 'filename': display_filename, 'file_url': attachment.file.url,
                'uploaded_at': attachment.uploaded_at.isoformat()
            })

        except ValueError as ve:
            errors.append({'filename': original_filename, 'error': str(ve)})
        except Exception as upload_err:
            logger.error(f"Falha inesperada no upload LEGADO do arquivo '{original_filename}': {upload_err}", exc_info=True)
            errors.append({'filename': original_filename, 'error': 'Erro interno durante o upload.'})
            if attachment and attachment.pk:
                try: attachment.delete()
                except Exception: pass

    # Define status final
    final_status_code = 200 if not errors else (207 if saved_attachments_info else 400)
    overall_status = 'failure'
    if saved_attachments_info and errors: overall_status = 'partial_success'
    elif saved_attachments_info: overall_status = 'success'
    response_message = f'{len(saved_attachments_info)} arquivo(s) salvos, {len(errors)} falha(s) (Upload Legado).'

    # Atualiza 'updated_at' do chat se algum arquivo foi salvo
    if last_upload_time:
        try:
            chat.updated_at = last_upload_time
            chat.save(update_fields=['updated_at'])
        except Exception as final_save_err:
            logger.error(f"Erro ao atualizar 'updated_at' final para chat {chat.id} após upload LEGADO: {final_save_err}")

    return JsonResponse({
        'status': overall_status, 'message': response_message,
        'saved_attachments': saved_attachments_info, 'upload_errors': errors
    }, status=final_status_code)


# ==============================================================================
# VIEW DE MONITORAMENTO DE CUSTOS
# ==============================================================================
@login_required
def api_cost_monitor_view(request):
    """Exibe um resumo dos custos de uso da API com filtros."""
    can_view_all = request.user.has_perm('ia.view_all_api_costs')

    form = None
    summary_data = {}
    user_to_filter = request.user if not can_view_all else None
    selected_start_date = None
    selected_end_date = None
    selected_model_name = None

    if ApiUsageFilterForm:
        form_initial = {}
        if user_to_filter:
            form_initial['user'] = user_to_filter

        form = ApiUsageFilterForm(request.GET or None, initial=form_initial)

        # Popula Choices do Modelo dinamicamente
        try:
            distinct_models = ApiUsageLog.objects.values_list('model_name', flat=True)\
                                                 .distinct().order_by('model_name')
            model_choices = [('', '-- Todos os Modelos --')] + [(name, name) for name in distinct_models if name]
            form.fields['model_name'].choices = model_choices
        except Exception as e:
            logger.error(f"Erro ao buscar modelos distintos para filtro: {e}")
            form.fields['model_name'].choices = [('', '-- Todos os Modelos --')]

        # Ajusta campo de usuário baseado na permissão
        if not can_view_all:
            if 'user' in form.fields:
                form.fields['user'].widget = forms.HiddenInput()
                form.fields['user'].required = False
                form.fields['user'].queryset = User.objects.filter(pk=request.user.pk)
                form.fields['user'].disabled = True
        # else: Campo fica visível e editável

        if form.is_valid():
            selected_start_date = form.cleaned_data.get('start_date')
            selected_end_date = form.cleaned_data.get('end_date')
            selected_model_name = form.cleaned_data.get('model_name')

            if can_view_all:
                user_to_filter = form.cleaned_data.get('user')
            # else: user_to_filter já está request.user

    else: # Fallback se ApiUsageFilterForm não existir
        form = lambda: None
        form.is_valid = lambda: False
        form.cleaned_data = {}
        form.as_p = lambda: "<p>Formulário de filtro indisponível.</p>"


    # Obtém Dados Sumarizados
    try:
        summary_data = ApiUsageLog.get_cost_summary(
            user=user_to_filter,
            start_date=selected_start_date,
            end_date=selected_end_date,
            model_name=selected_model_name
        )
    except Exception as summary_error:
        logger.error(f"Erro ao calcular resumo de custos: {summary_error}", exc_info=True)


    context = {
        'summary': summary_data,
        'form': form,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_user_filter': user_to_filter,
        'selected_model_filter': selected_model_name,
        'can_view_all': can_view_all,
    }
    return render(request, 'ia/api_cost_monitor.html', context)